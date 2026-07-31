"""
Excel loading and validation for Euler Mail.
Reads the recipient list, validates email column presence and format.
"""
import re
import logging
from pathlib import Path
from typing import Optional, Tuple, List

from openpyxl import load_workbook

from euler_mail.data.models import ValidationError

logger = logging.getLogger(__name__)

EMAIL_REGEX = re.compile(
    r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$'
)

# All accepted names for the email column (case-insensitive match)
MAIL_COLUMN_ALIASES = {
    "mail", "email", "e-mail", "emailaddress", "email address",
    "email_address", "recipient", "recipient_email",
}


def get_mail_column(headers: List[str]) -> Optional[str]:
    """Return the original header string for the email column, or None."""
    for h in headers:
        if h.strip().lower().replace(" ", "") in MAIL_COLUMN_ALIASES or \
           h.strip().lower() in MAIL_COLUMN_ALIASES:
            return h
    return None


def load_excel(
    path: "str | Path",
) -> Tuple[List[str], List[dict], List[ValidationError]]:
    """
    Load an Excel file (.xlsx/.xls) and return:
        headers       — list of column header strings (row 1)
        rows          — list of dicts, one per data row
        errors        — list of ValidationError (never raises; caller decides)

    A missing or empty mail column is reported as a ValidationError.
    """
    path = Path(path)
    validation_errors: List[ValidationError] = []

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        return [], [], [ValidationError(0, "", f"Cannot open file: {exc}")]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # ── Headers ──────────────────────────────────────────────────────────────
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [], [ValidationError(0, "", "The spreadsheet appears to be empty.")]

    headers: List[str] = []
    for i, cell in enumerate(header_row):
        val = str(cell).strip() if cell is not None else f"Column_{i + 1}"
        headers.append(val)

    # Remove trailing empty headers
    while headers and headers[-1].startswith("Column_"):
        headers.pop()

    # ── Identify email column ──────────────────────────────────────────────
    mail_col = get_mail_column(headers)
    if mail_col is None:
        validation_errors.append(
            ValidationError(
                0, "",
                f"No email column found. Expected one of: "
                f"{', '.join(sorted(MAIL_COLUMN_ALIASES))}."
            )
        )

    # ── Data rows ─────────────────────────────────────────────────────────
    rows: List[dict] = []
    for row_idx, raw_row in enumerate(rows_iter, start=2):
        row_dict: dict = {}
        for col_idx, (header, value) in enumerate(zip(headers, raw_row)):
            row_dict[header] = str(value).strip() if value is not None else ""

        # Fill any missing columns
        for header in headers:
            row_dict.setdefault(header, "")

        # Validate email
        if mail_col:
            email_val = row_dict.get(mail_col, "").strip()
            if not email_val:
                validation_errors.append(
                    ValidationError(row_idx, "", f"Row {row_idx}: email cell is empty.")
                )
            elif not EMAIL_REGEX.match(email_val):
                validation_errors.append(
                    ValidationError(
                        row_idx, email_val,
                        f"Row {row_idx}: '{email_val}' is not a valid email address."
                    )
                )

        rows.append(row_dict)

    wb.close()
    logger.info(
        f"Loaded {len(rows)} rows from '{path.name}'. "
        f"Errors: {len(validation_errors)}"
    )
    return headers, rows, validation_errors
